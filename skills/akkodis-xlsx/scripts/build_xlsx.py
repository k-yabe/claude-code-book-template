#!/usr/bin/env python3
"""AKKODiS ブランド準拠 XLSX を生成する。

入力 (Markdown ライク):
    # タイトル: ...
    ## シート名
    | 列1 | 列2 |
    | --- | --- |
    | 値  | 値  |

    ## シート名 [chart]   ← 末尾に [chart] を付けると自動で棒グラフを追加

依存: openpyxl
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).resolve().parent))
from notation import correct  # noqa: E402

NAVY = "FF001F33"
NAVY_RGB = "001F33"
GOLD = "FFFFB81C"
WHITE = "FFFFFFFF"
ZEBRA = "FFF4F6F8"
WARN = "FFF4C7C3"
GRID = "FFE0E4E8"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Noto Sans JP", size=11, bold=True, color=WHITE)
BODY_FONT = Font(name="Noto Sans JP", size=11, color="FF000000")
TITLE_FONT = Font(name="Noto Sans JP", size=16, bold=True, color=NAVY)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
GOLD_FILL = PatternFill("solid", fgColor=GOLD)
WARN_FILL = PatternFill("solid", fgColor=WARN)

THIN = Side(border_style="thin", color=GRID)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass
class Sheet:
    name: str
    header: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)
    chart: bool = False


@dataclass
class Book:
    title: str = "Untitled"
    sheets: list[Sheet] = field(default_factory=list)


def parse_input(text: str) -> Book:
    book = Book()
    current: Sheet | None = None
    table_buffer: list[str] = []
    in_table = False

    def split_cells(line: str) -> list[str]:
        line = line.strip()
        if line.startswith("|"):
            line = line[1:]
        if line.endswith("|"):
            line = line[:-1]
        return [c.strip() for c in line.split("|")]

    def flush_table() -> None:
        nonlocal table_buffer, in_table, current
        if not table_buffer or current is None:
            table_buffer = []
            in_table = False
            return
        rows = [r for r in table_buffer if r.strip()]
        table_buffer = []
        in_table = False
        if len(rows) < 2:
            return
        header = split_cells(rows[0])
        body_rows = rows[2:] if re.match(r"^\s*\|?[\s\-:|]+\|?\s*$", rows[1]) else rows[1:]
        current.header = [correct(h) for h in header]
        current.rows = [[correct(c) for c in split_cells(r)] for r in body_rows]

    for raw in text.splitlines():
        line = raw.rstrip()
        m = re.match(r"^#\s*(?:タイトル|title)\s*[:：]\s*(.+)$", line, re.IGNORECASE)
        if m:
            book.title = correct(m.group(1).strip())
            continue
        m = re.match(r"^##\s+(.+)$", line)
        if m:
            flush_table()
            raw_name = m.group(1).strip()
            chart_flag = bool(re.search(r"\[chart\]\s*$", raw_name, re.IGNORECASE))
            name = re.sub(r"\s*\[chart\]\s*$", "", raw_name, flags=re.IGNORECASE)
            current = Sheet(name=correct(name)[:31], chart=chart_flag)
            book.sheets.append(current)
            continue
        if line.lstrip().startswith("|"):
            in_table = True
            table_buffer.append(line)
            continue
        if in_table and not line.strip():
            flush_table()
            continue
    flush_table()
    return book


def _is_number(v: str) -> bool:
    if v is None:
        return False
    s = v.replace(",", "").replace("¥", "").replace("円", "").strip()
    if s.endswith("%"):
        s = s[:-1]
    try:
        float(s)
        return True
    except ValueError:
        return False


def _to_number(v: str) -> float:
    s = v.replace(",", "").replace("¥", "").replace("円", "").strip()
    if s.endswith("%"):
        return float(s[:-1]) / 100.0
    return float(s)


def _is_pct_column(header: str, values: list[str]) -> bool:
    if "率" in header or "%" in header:
        return True
    pct_count = sum(1 for v in values if isinstance(v, str) and v.strip().endswith("%"))
    return pct_count > 0 and pct_count >= len(values) * 0.5


def write_sheet(ws, sheet: Sheet, *, is_first: bool, title: str) -> None:
    start_row = 1
    if is_first and title:
        ws.cell(row=1, column=1, value=correct(title)).font = TITLE_FONT
        ws.row_dimensions[1].height = 24
        start_row = 3

    if not sheet.header:
        return

    # ヘッダー
    for c, h in enumerate(sheet.header, start=1):
        cell = ws.cell(row=start_row, column=c, value=correct(h))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[start_row].height = 24

    # 列ごとの判定
    pct_cols: set[int] = set()
    for c, h in enumerate(sheet.header, start=1):
        col_values = [row[c - 1] for row in sheet.rows if c - 1 < len(row)]
        if _is_pct_column(h, col_values):
            pct_cols.add(c)

    # ボディ
    for r_offset, row in enumerate(sheet.rows):
        r = start_row + 1 + r_offset
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if _is_number(v):
                cell.value = _to_number(v)
                if c in pct_cols or v.strip().endswith("%"):
                    cell.number_format = "0%"
                else:
                    cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            # ゼブラ（条件付き書式は別途）
            if r_offset % 2 == 1:
                cell.fill = ZEBRA_FILL

    # オートフィルタ
    last_col = get_column_letter(len(sheet.header))
    last_row = start_row + len(sheet.rows)
    if sheet.rows:
        ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"

    # 列幅
    for c, h in enumerate(sheet.header, start=1):
        max_len = len(str(h))
        for row in sheet.rows:
            if c - 1 < len(row):
                max_len = max(max_len, len(str(row[c - 1])))
        ws.column_dimensions[get_column_letter(c)].width = min(40, max(12, max_len + 2))

    # フリーズペイン: ヘッダー直下＆1列目
    ws.freeze_panes = ws.cell(row=start_row + 1, column=2)

    # シートタブ色
    ws.sheet_properties.tabColor = NAVY_RGB

    # 条件付き書式: % 列は閾値で色分け
    if sheet.rows and pct_cols:
        first_data_row = start_row + 1
        last_data_row = start_row + len(sheet.rows)
        for c in pct_cols:
            col = get_column_letter(c)
            rng = f"{col}{first_data_row}:{col}{last_data_row}"
            # >= 100% → Gold
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["1"],
                    fill=PatternFill("solid", fgColor=GOLD),
                    font=Font(name="Noto Sans JP", size=11, bold=True, color="FF001F33"),
                ),
            )
            # < 80% → 薄赤
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="lessThan",
                    formula=["0.8"],
                    fill=PatternFill("solid", fgColor=WARN),
                ),
            )

    # 印刷設定
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{start_row}:{start_row}"
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6, header=0.3, footer=0.3)
    ws.oddHeader.center.text = correct(title or sheet.name)
    ws.oddHeader.center.size = 10
    ws.oddHeader.center.color = NAVY_RGB
    ws.oddFooter.right.text = "&P / &N"
    ws.oddFooter.right.size = 9
    ws.oddFooter.left.text = "AKKODiS Internal Use"
    ws.oddFooter.left.size = 9
    ws.oddFooter.left.color = "5A6470"

    # チャート（指定がある場合）
    if sheet.chart:
        _add_chart(ws, sheet, start_row=start_row)


def _add_chart(ws, sheet: Sheet, *, start_row: int) -> None:
    """先頭の文字列列をカテゴリ、最初の数値列を値として棒グラフを追加する。"""
    if not sheet.rows:
        return
    # 数値列を探す
    numeric_col = None
    for c, h in enumerate(sheet.header, start=1):
        col_values = [row[c - 1] for row in sheet.rows if c - 1 < len(row)]
        if col_values and all(_is_number(v) for v in col_values):
            numeric_col = c
            break
    if numeric_col is None or numeric_col == 1:
        return  # ラベル列が無い

    last_data_row = start_row + len(sheet.rows)
    chart = BarChart()
    chart.type = "col"
    chart.style = 2
    chart.title = correct(sheet.header[numeric_col - 1])
    chart.y_axis.title = correct(sheet.header[numeric_col - 1])
    chart.x_axis.title = correct(sheet.header[0])
    chart.legend = None
    chart.height = 8
    chart.width = 16
    data = Reference(
        ws,
        min_col=numeric_col,
        min_row=start_row,
        max_row=last_data_row,
        max_col=numeric_col,
    )
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=last_data_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.dLbls = DataLabelList(showVal=True)
    # 配置: 表の右側
    anchor_col = get_column_letter(len(sheet.header) + 2)
    ws.add_chart(chart, f"{anchor_col}{start_row}")


def build(book: Book) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for i, sheet in enumerate(book.sheets):
        ws = wb.create_sheet(title=sheet.name)
        write_sheet(ws, sheet, is_first=(i == 0), title=book.title)
    if not book.sheets:
        ws = wb.create_sheet(title="Summary")
        ws["A1"] = correct(book.title)
        ws["A1"].font = TITLE_FONT
    return wb


def main(argv: list[str]) -> int:
    ap = argparse.ArgumentParser(description="AKKODiS ブランド準拠 XLSX 生成")
    ap.add_argument("--input", type=Path, required=True)
    ap.add_argument("--output", type=Path, required=True)
    args = ap.parse_args(argv)
    book = parse_input(args.input.read_text(encoding="utf-8"))
    wb = build(book)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(args.output))
    print(f"[ok] {args.output}  ({len(book.sheets)} sheets)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
